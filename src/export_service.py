#!/usr/bin/env python3
"""
Export Service for HR Tech Lead Generation System
Handles export to CSV, Excel with beautiful formatting
"""

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import Opportunity, SIGNAL_TYPES

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting opportunities to various formats"""

    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def export_to_csv(
        self,
        opportunities: List[Opportunity],
        filename: Optional[str] = None,
        include_content: bool = False,
    ) -> str:
        """Export opportunities to CSV file"""
        if not opportunities:
            logger.warning("No opportunities to export")
            return ""

        if filename is None:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"opportunities_{timestamp}.csv"

        filepath = self.output_dir / filename

        # Prepare data
        data = []
        for opp in opportunities:
            row = {
                "Title": opp.title,
                "Company": opp.company,
                "Person": opp.person,
                "Email": opp.email,
                "URL": opp.url,
                "Date": opp.date,
                "Relevance Score": f"{opp.relevance_score:.2f}",
                "Signal Type": SIGNAL_TYPES.get(opp.signal_type, f"Signal {opp.signal_type}"),
                "Source": opp.source,
            }
            if include_content:
                row["Content"] = opp.content[:500]  # Truncate for CSV
            data.append(row)

        # Create DataFrame and export
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compatibility

        logger.info(f"✅ Exported {len(opportunities)} opportunities to {filepath}")
        return str(filepath)

    def export_to_excel(
        self,
        opportunities: List[Opportunity],
        filename: Optional[str] = None,
        include_content: bool = False,
        apply_formatting: bool = True,
    ) -> str:
        """Export opportunities to Excel with beautiful formatting"""
        if not opportunities:
            logger.warning("No opportunities to export")
            return ""

        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, falling back to CSV")
            return self.export_to_csv(opportunities, filename, include_content)

        if filename is None:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"opportunities_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Prepare data
        data = []
        for opp in opportunities:
            row = {
                "Title": opp.title,
                "Company": opp.company,
                "Person": opp.person,
                "Email": opp.email,
                "URL": opp.url,
                "Date": opp.date,
                "Relevance Score": opp.relevance_score,
                "Signal Type": SIGNAL_TYPES.get(opp.signal_type, f"Signal {opp.signal_type}"),
                "Source": opp.source,
            }
            if include_content:
                row["Content"] = opp.content[:1000]  # Truncate for Excel
            data.append(row)

        # Create DataFrame
        df = pd.DataFrame(data)

        # Export to Excel
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Opportunities", index=False)

            if apply_formatting:
                self._apply_excel_formatting(writer, df)

        logger.info(f"✅ Exported {len(opportunities)} opportunities to {filepath}")
        return str(filepath)

    def _apply_excel_formatting(self, writer, df: pd.DataFrame) -> None:
        """Apply beautiful formatting to Excel file"""
        workbook = writer.book
        worksheet = writer.sheets["Opportunities"]

        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col)),
            )
            # Set reasonable limits
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

        # Format relevance score column (highlight high scores)
        relevance_col = df.columns.get_loc("Relevance Score") + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=relevance_col)
            score = cell.value
            if isinstance(score, (int, float)):
                if score >= 0.8:
                    cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )  # Light green
                elif score >= 0.7:
                    cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )  # Light yellow
                cell.number_format = "0.00"

        # Format email column (make it clickable)
        email_col = df.columns.get_loc("Email") + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=email_col)
            if cell.value and cell.value != "Manual validation needed":
                cell.font = Font(color="0563C1", underline="single")  # Blue, underlined

        # Format URL column (make it clickable)
        url_col = df.columns.get_loc("URL") + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=url_col)
            if cell.value:
                cell.font = Font(color="0563C1", underline="single")  # Blue, underlined
                cell.hyperlink = cell.value

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Add filter
        worksheet.auto_filter.ref = worksheet.dimensions

    def export_summary_report(
        self,
        opportunities: List[Opportunity],
        metrics: dict,
        filename: Optional[str] = None,
    ) -> str:
        """Export summary report with metrics"""
        if filename is None:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"summary_report_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, cannot create summary report")
            return ""

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Summary sheet
            summary_data = {
                "Metric": [
                    "Total Opportunities",
                    "Average Relevance Score",
                    "High Quality Count (≥0.8)",
                    "Quality Percentage",
                    "Emails Found",
                    "Email Found Percentage",
                    "Recent Opportunities (30 days)",
                    "Recent Percentage",
                ],
                "Value": [
                    metrics.get("total_opportunities", 0),
                    f"{metrics.get('average_relevance_score', 0):.3f}",
                    metrics.get("high_quality_count", 0),
                    f"{metrics.get('quality_percentage', 0):.1f}%",
                    metrics.get("email_found_count", 0),
                    f"{metrics.get('email_found_percentage', 0):.1f}%",
                    metrics.get("recent_opportunities", 0),
                    f"{metrics.get('recent_percentage', 0):.1f}%",
                ],
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)

            # Opportunities sheet
            self.export_to_excel(opportunities, filename=None, include_content=False, apply_formatting=False)
            # Reuse the same writer but write to different sheet
            opp_data = [opp.to_dict() for opp in opportunities]
            df_opps = pd.DataFrame(opp_data)
            df_opps.to_excel(writer, sheet_name="Opportunities", index=False)

            # Distribution by signal
            if "signals_distribution" in metrics:
                signal_data = {
                    "Signal Type": [
                        SIGNAL_TYPES.get(sid, f"Signal {sid}")
                        for sid in metrics["signals_distribution"].keys()
                    ],
                    "Count": list(metrics["signals_distribution"].values()),
                }
                df_signals = pd.DataFrame(signal_data)
                df_signals.to_excel(writer, sheet_name="By Signal", index=False)

        logger.info(f"✅ Exported summary report to {filepath}")
        return str(filepath)

